#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SQL Server 数据库字典生成工具

根据模板格式，为每个业务库生成一个 Excel 数据字典文件。
每个 Sheet 代表一个 Schema，包含该 Schema 下所有表的字段信息。
"""

import os
import sys
import pyodbc
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import DB_CONFIG, EXCLUDE_DB_LIST, OUTPUT_DIR


# ============================================================
# Excel 样式定义（参照模板风格）
# ============================================================

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=12, bold=True)
LABEL_FONT = Font(name="微软雅黑", size=11, bold=True)
DATA_FONT = Font(name="微软雅黑", size=10)
TABLE_NAME_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
NOTE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 模板列头（8列）
COLUMN_HEADERS = [ "列名", "列类型", "长度", "精度", "是否为空", "默认值", "列名备注", "是否为主键"]

# 列宽设置（8列）
COLUMN_WIDTHS = {
    1: 22,   # 列名
    2: 16,   # 列类型
    3: 10,   # 长度
    4: 10,   # 精度
    5: 10,   # 是否为空
    6: 16,   # 默认值
    7: 32,   # 列名备注
    8: 12,  # 是否为主键
}


def apply_style(worksheet, row, col, value, font=None, fill=None, alignment=None, border=None):
    """设置单元格样式"""
    cell = worksheet.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def create_workbook_template(workbook, db_name):
    """创建符合模板格式的文件头"""
    ws = workbook.create_sheet(title=db_name)

    # Row 1: 库名
    apply_style(ws, 1, 1, "库名", LABEL_FONT, NOTE_FILL, Alignment(horizontal="right"), THIN_BORDER)
    apply_style(ws, 1, 2, db_name, TITLE_FONT, None, Alignment(horizontal="left"), THIN_BORDER)
    # 合并 B1:H1（8列，最后一列是H）
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=8)

    # Row 2: 备注
    apply_style(ws, 2, 1, "备注", LABEL_FONT, NOTE_FILL, Alignment(horizontal="right"), THIN_BORDER)
    apply_style(ws, 2, 2, "", border=THIN_BORDER)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=8)

    # Row 3: 列头
    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        apply_style(
            ws, 3, col_idx, header, HEADER_FONT, HEADER_FILL,
            Alignment(horizontal="center", vertical="center"),
            THIN_BORDER,
        )

    # 设置列宽
    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return ws


def fetch_database_list(cursor):
    """获取所有用户数据库列表"""
    cursor.execute(
        """
        SELECT name FROM sys.databases
        WHERE database_id > 4
          AND name NOT IN ({})
          AND state_desc = 'ONLINE'
        ORDER BY name
        """.format(",".join(["?" for _ in EXCLUDE_DB_LIST])),
        EXCLUDE_DB_LIST,
    )
    return [row[0] for row in cursor.fetchall()]


def fetch_schema_tables(cursor):
    """获取当前库中所有 Schema 和表信息"""
    cursor.execute(
        """
        SELECT
            s.name AS schema_name,
            t.name AS table_name,
            CAST(ep.value AS NVARCHAR(500)) AS table_desc
        FROM sys.tables t
        JOIN sys.schemas s ON t.schema_id = s.schema_id
        LEFT JOIN sys.extended_properties ep
            ON ep.major_id = t.object_id AND ep.minor_id = 0 AND ep.name = 'MS_Description'
        ORDER BY s.name, t.name
        """
    )

    # Group by schema
    result = {}
    for row in cursor.fetchall():
        schema = row.schema_name
        if schema not in result:
            result[schema] = []
        result[schema].append({
            "schema": schema,
            "table_name": row.table_name,
            "table_desc": row.table_desc or "",
        })
    return result


def fetch_columns(cursor, schema_name, table_name):
    """获取表的列信息（长度与精度分开返回）"""
    cursor.execute(
        """
        SELECT
            c.name AS column_name,
            TYPE_NAME(c.user_type_id) AS data_type,
            -- 长度
            CASE
                WHEN TYPE_NAME(c.user_type_id) IN ('varchar', 'nvarchar', 'char', 'nchar', 'varbinary',
                                                    'binary', 'text', 'ntext', 'image')
                    THEN CASE WHEN c.max_length = -1 THEN '(MAX)' ELSE CAST(c.max_length AS VARCHAR(10)) END
                WHEN TYPE_NAME(c.user_type_id) IN ('numeric', 'decimal')
                    THEN CAST(c.precision AS VARCHAR(10))
                ELSE ''
            END AS col_length,
            -- 精度
            CASE
                WHEN TYPE_NAME(c.user_type_id) IN ('numeric', 'decimal')
                    THEN CAST(c.scale AS VARCHAR(10))
                WHEN TYPE_NAME(c.user_type_id) IN ('datetime2', 'datetimeoffset', 'time')
                    THEN CAST(c.scale AS VARCHAR(10))
                WHEN TYPE_NAME(c.user_type_id) IN ('float', 'real')
                    THEN CAST(c.precision AS VARCHAR(10))
                ELSE ''
            END AS col_precision,
            c.is_nullable,
            ISNULL(defn.definition, '') AS default_value,
            ISNULL(ep.value, '') AS column_desc
        FROM sys.columns c
        LEFT JOIN sys.default_constraints defn
            ON defn.parent_object_id = c.object_id AND defn.parent_column_id = c.column_id
        LEFT JOIN sys.extended_properties ep
            ON ep.major_id = c.object_id AND ep.minor_id = c.column_id AND ep.name = 'MS_Description'
        WHERE c.object_id = OBJECT_ID(QUOTENAME(?) + '.' + QUOTENAME(?))
        ORDER BY c.column_id
        """,
        schema_name, table_name,
    )

    columns = []
    for row in cursor.fetchall():
        columns.append({
            "column_name": row.column_name,
            "data_type": row.data_type,
            "length": str(row.col_length) if row.col_length else "",
            "precision": str(row.col_precision) if row.col_precision else "",
            "is_nullable": row.is_nullable,
            "default_value": str(row.default_value) if row.default_value else "",
            "column_desc": str(row.column_desc) if row.column_desc else "",
        })
    return columns


def fetch_primary_keys(cursor, schema_name, table_name):
    """获取表的主键列名集合"""
    cursor.execute(
        """
        SELECT c.name
        FROM sys.indexes i
        JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id
        JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
        WHERE i.object_id = OBJECT_ID(QUOTENAME(?) + '.' + QUOTENAME(?))
          AND i.is_primary_key = 1
        ORDER BY ic.key_ordinal
        """,
        schema_name, table_name,
    )
    return {row[0] for row in cursor.fetchall()}


def fetch_indexes(cursor, schema_name, table_name):
    """获取表的索引信息"""
    cursor.execute(
        """
        SELECT
            i.name AS index_name,
            i.is_unique,
            STRING_AGG(c.name, ',') WITHIN GROUP (ORDER BY ic.key_ordinal) AS columns,
            i.type_desc
        FROM sys.indexes i
        JOIN sys.index_columns ic ON ic.object_id = i.object_id AND ic.index_id = i.index_id
        JOIN sys.columns c ON c.object_id = ic.object_id AND c.column_id = ic.column_id
        WHERE i.object_id = OBJECT_ID(QUOTENAME(?) + '.' + QUOTENAME(?))
          AND i.is_primary_key = 0
          AND i.type > 0  -- exclude heaps
        GROUP BY i.name, i.is_unique, i.type_desc
        ORDER BY i.name
        """,
        schema_name, table_name,
    )

    indexes = []
    for row in cursor.fetchall():
        indexes.append({
            "index_name": row.index_name,
            "is_unique": row.is_unique,
            "columns": row.columns,
            "type_desc": row.type_desc,
        })
    return indexes


def write_table_section(worksheet, table_info, columns, pk_set, start_row):
    """写入一个表的分区（表名行 + 表备注行 + 列头行 + 数据行），返回下一区域起始行号"""
    full_table_name = f"{table_info['schema']}.{table_info['table_name']}"
    table_desc = table_info.get("table_desc", "")
    max_col = len(COLUMN_HEADERS)  # 8 列，最后一列是 H

    # ── 表名行 ──
    apply_style(
        worksheet, start_row, 1, "表名", LABEL_FONT, TABLE_NAME_FILL,
        Alignment(horizontal="right"), THIN_BORDER,
    )
    apply_style(
        worksheet, start_row, 2, full_table_name, TITLE_FONT, TABLE_NAME_FILL,
        Alignment(horizontal="left"), THIN_BORDER,
    )
    worksheet.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=max_col)
    start_row += 1

    # ── 表备注行 ──
    apply_style(
        worksheet, start_row, 1, "表备注", LABEL_FONT, NOTE_FILL,
        Alignment(horizontal="right"), THIN_BORDER,
    )
    apply_style(
        worksheet, start_row, 2, table_desc, None, NOTE_FILL,
        Alignment(horizontal="left"), THIN_BORDER,
    )
    worksheet.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=max_col)
    start_row += 1

    # ── 列头行 ──
    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        apply_style(
            worksheet, start_row, col_idx, header, HEADER_FONT, HEADER_FILL,
            Alignment(horizontal="center", vertical="center"),
            THIN_BORDER,
        )
    start_row += 1

    # ── 字段数据行 ──
    for col_data in columns:
        is_pk = "是" if col_data["column_name"] in pk_set else ""
        nullable = "是" if col_data["is_nullable"] else "否"

        # 不在每行显示表名和表备注，只保留边框
        #apply_style(worksheet, start_row, 1, None, DATA_FONT, border=THIN_BORDER)
        #apply_style(worksheet, start_row, 2, None, DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 1, col_data["column_name"], DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 2, col_data["data_type"], DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 3, col_data["length"], DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 4, col_data["precision"], DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 5, nullable, DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 6, col_data["default_value"], DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 7, col_data["column_desc"], DATA_FONT, border=THIN_BORDER)
        apply_style(worksheet, start_row, 8, is_pk, DATA_FONT, border=THIN_BORDER)
        start_row += 1

    return start_row

def generate_database_dict(db_name):
    """为单个数据库生成数据字典 Excel 文件"""
    print(f"  正在处理数据库: {db_name} ...")

    conn = pyodbc.connect(
        server=DB_CONFIG["server"],
        uid=DB_CONFIG["uid"],
        pwd=DB_CONFIG["pwd"],
        database=db_name,
        driver=DB_CONFIG["driver"],
        timeout=30,
    )
    cursor = conn.cursor()

    # 获取表信息
    schema_tables = fetch_schema_tables(cursor)

    if not schema_tables:
        print(f"    警告: 数据库 '{db_name}' 中没有找到用户表")
        cursor.close()
        conn.close()
        return None

    # 创建 Excel 工作簿（空工作簿，之后用 create_sheet 创建 Sheet）
    wb = openpyxl.Workbook()
    # 删除默认空 sheet
    if wb.sheetnames:
        wb.remove(wb[wb.sheetnames[0]])

    total_tables = sum(len(tables) for tables in schema_tables.values())
    print(f"    共 {total_tables} 个表, {len(schema_tables)} 个 Schema")

    for schema_name, tables in schema_tables.items():
        ws = create_workbook_template(wb, schema_name)

        # 跳过库名/备注行后的起始行
        current_row = 4

        for table_info in tables:
            full_name = f"{table_info['schema']}.{table_info['table_name']}"

            # 获取列信息
            columns = fetch_columns(cursor, table_info["schema"], table_info["table_name"])

            # 获取主键信息
            pk_set = fetch_primary_keys(cursor, table_info["schema"], table_info["table_name"])

            # 写入表区域
            current_row = write_table_section(ws, table_info, columns, pk_set, current_row)

            # 每个表之间空一行
            current_row += 1

        print(f"    Schema '{schema_name}': {len(tables)} 个表 -> Sheet '{schema_name}'")

    cursor.close()
    conn.close()

    # 保存文件
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    file_path = os.path.join(OUTPUT_DIR, f"{db_name}_数据字典.xlsx")
    wb.save(file_path)
    print(f"  [OK] 已生成: {file_path}")
    return file_path


def main():
    print("=" * 60)
    print("SQL Server 数据库字典生成工具")
    print(f"服务器: {DB_CONFIG['server']}")
    print(f"排除系统库: {', '.join(EXCLUDE_DB_LIST)}")
    print(f"输出目录: {os.path.abspath(OUTPUT_DIR)}")
    print("=" * 60)

    # 连接到 master 获取数据库列表
    print("\n正在连接数据库服务器...")
    try:
        master_conn = pyodbc.connect(
            server=DB_CONFIG["server"],
            uid=DB_CONFIG["uid"],
            pwd=DB_CONFIG["pwd"],
            database="master",
            driver=DB_CONFIG["driver"],
            timeout=10,
        )
    except pyodbc.Error as e:
        print(f"\n[错误] 连接失败: {e}")
        sys.exit(1)

    master_cursor = master_conn.cursor()
    db_list = fetch_database_list(master_cursor)
    master_cursor.close()
    master_conn.close()

    if not db_list:
        print("没有找到业务数据库!")
        sys.exit(0)

    print(f"\n找到 {len(db_list)} 个业务库: {', '.join(db_list)}\n")

    # 为每个库生成字典
    generated_files = []
    for db_name in db_list:
        try:
            file_path = generate_database_dict(db_name)
            if file_path:
                generated_files.append(file_path)
        except Exception as e:
            print(f"\n[错误] 数据库 '{db_name}' 处理失败: {e}")
            import traceback
            traceback.print_exc()

    print("\n" + "=" * 60)
    print(f"生成完成! 共生成 {len(generated_files)} 个文件:")
    for f in generated_files:
        print(f"  {f}")
    print("=" * 60)


if __name__ == "__main__":
    main()
